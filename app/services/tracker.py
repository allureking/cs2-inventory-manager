"""
每日收益追踪服务

功能：
  snapshot_daily()       — 自动抓取当日数据并写入 daily_tracker
  get_daily_records()    — 查询日期范围记录
  get_monthly_summary()  — 月度汇总（动态聚合）
  update_record()        — 手动编辑某天数据
  import_from_excel()    — 从 xlsx 导入历史数据
  export_to_excel()      — 导出 xlsx
"""

from __future__ import annotations

import io
import logging
import re
from calendar import monthrange
from datetime import date, datetime, timezone
from typing import Optional
from zoneinfo import ZoneInfo

from sqlalchemy import and_, select, func, or_
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import ACTIVE_STATUSES
from app.models.db_models import DailyTracker, InventoryItem, PriceSnapshot, TrackerConfig
from app.services.pricing import get_latest_prices

logger = logging.getLogger(__name__)

# ── 年化参数 ──────────────────────────────────────────────────────────────
# 公式：期望周期 = R + (1-S)×CD，有效天数 = (365/期望周期)×R
# R=租期, S=转租成功率, CD=冷却天数

# 非大会员：传统模式（无转租，S=0）
_FEE_RATE = 0.20        # 服务费 20%
_NET_RATE = 0.80        # 净比例
_SHORT_DAYS = 182.5     # R=8, CD=8 → 周期16 → (365/16)*8 = 182.5
_LONG_DAYS = 267.7      # R=22, CD=8 → 周期30 → (365/30)*22 = 267.7

# 大会员 V3：0CD 转租（S=0.85, CD=8, 服务费 10%）
_VIP_FEE_RATE = 0.10    # 服务费 10%
_VIP_NET_RATE = 0.90    # 净比例
_VIP_SHORT_DAYS = 317.4  # R=8, 期望周期=8+0.15*8=9.2 → (365/9.2)*8
_VIP_LONG_DAYS = 346.1   # R=22, 期望周期=22+0.15*8=23.2 → (365/23.2)*22

# 综合年化权重
_SHORT_WEIGHT = 0.7
_LONG_WEIGHT = 0.3



# ══════════════════════════════════════════════════════════════
#  解析 & 计算
# ══════════════════════════════════════════════════════════════

def _parse_stats_desc(text: str) -> dict:
    """
    解析悠悠租出统计摘要。
    输入: "件数：3,320｜价值：¥3612634.69｜总租金：¥2466.82/天"
    输出: {"count": 3320, "value": 3612634.69, "income": 2466.82}
    """
    result = {"count": 0, "value": 0.0, "income": 0.0}
    if not text:
        return result

    # 件数
    m = re.search(r"件数[：:]\s*([\d,]+)", text)
    if m:
        result["count"] = int(m.group(1).replace(",", ""))

    # 价值
    m = re.search(r"价值[：:]\s*¥?([\d,.]+)", text)
    if m:
        result["value"] = float(m.group(1).replace(",", ""))

    # 总租金
    m = re.search(r"总租金[：:]\s*¥?([\d,.]+)", text)
    if m:
        result["income"] = float(m.group(1).replace(",", ""))

    return result


def _calc_annuals(daily_income: float, rented_value: float, is_vip: bool = True) -> dict:
    """计算短租/长租/综合年化收益率（根据是否大会员选取参数）"""
    if not rented_value or rented_value <= 0:
        return {"short": None, "long": None, "combined": None}

    net_rate = _VIP_NET_RATE if is_vip else _NET_RATE
    short_days = _VIP_SHORT_DAYS if is_vip else _SHORT_DAYS
    long_days = _VIP_LONG_DAYS if is_vip else _LONG_DAYS

    net = daily_income * net_rate
    short = net * short_days / rented_value
    long = net * long_days / rented_value
    combined = short * _SHORT_WEIGHT + long * _LONG_WEIGHT

    return {
        "short": round(short, 8),
        "long": round(long, 8),
        "combined": round(combined, 8),
    }


# ══════════════════════════════════════════════════════════════
#  自动快照
# ══════════════════════════════════════════════════════════════

async def snapshot_daily(is_vip: bool = True) -> dict:
    """
    抓取当日数据并写入 daily_tracker。
    """
    import time as _time
    t_start = _time.monotonic()
    from app.core.database import AsyncSessionLocal
    from app.services.youpin import fetch_lease_records, get_active_token

    today = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%Y-%m-%d")
    logger.info("snapshot_daily: START — date=%s", today)

    # 1) 租赁统计
    rental = {"count": 0, "value": 0.0, "income": 0.0}
    token = get_active_token()
    if token:
        try:
            _, _, stats_desc = await fetch_lease_records(page=1, page_size=1)
            rental = _parse_stats_desc(stats_desc)
        except Exception as e:
            logger.warning("tracker snapshot_daily: 获取租赁统计失败: %s", e)
    else:
        logger.warning("tracker snapshot_daily: 无悠悠 Token，跳过租赁数据")

    # 2) 库存数据
    async with AsyncSessionLocal() as db:
        # 总活跃件数（in_steam + rented_out）
        total_inv = (await db.execute(
            select(func.count(InventoryItem.id))
            .where(InventoryItem.status.in_(ACTIVE_STATUSES))
        )).scalar() or 0

        # in_steam 物品市值：直接使用悠悠 API 估值（与同步拉取一致）
        in_steam_value = 0.0
        if token:
            try:
                from app.services.youpin import fetch_stock_records
                _, _, stock_valuation = await fetch_stock_records(page=1, page_size=1)
                in_steam_value = float(str(stock_valuation).replace(",", "")) if stock_valuation else 0.0
            except Exception as e:
                logger.warning("tracker snapshot_daily: 获取库存估值失败: %s", e)

        # fallback: 如果悠悠 API 不可用，从 price_snapshot 计算
        if in_steam_value == 0.0:
            steam_name_rows = (await db.execute(
                select(InventoryItem.market_hash_name, func.count(InventoryItem.id).label("cnt"))
                .where(InventoryItem.status == "in_steam")
                .group_by(InventoryItem.market_hash_name)
            )).all()
            price_map = await get_latest_prices([r[0] for r in steam_name_rows], db)
            for name, cnt in steam_name_rows:
                p = price_map.get(name)
                if p:
                    in_steam_value += p * cnt

        # 库存总价值 = 悠悠 API 出租价值 + 悠悠 API 库存估值
        inv_value = rental["value"] + in_steam_value

        # 3) 涨跌：(当日市值 - 成本基准) / 成本基准
        # 成本基准继承前一天的值，用户可手动修改
        prev_row = (await db.execute(
            select(DailyTracker.cost_basis)
            .where(DailyTracker.date < today, DailyTracker.cost_basis.isnot(None))
            .order_by(DailyTracker.date.desc())
            .limit(1)
        )).scalar()
        cost_basis = prev_row if prev_row else _DEFAULT_COST_BASIS
        price_change = (inv_value - cost_basis) / cost_basis if cost_basis > 0 else 0.0

        # 4) 计算年化（根据大会员状态选取参数）
        annuals = _calc_annuals(rental["income"], rental["value"], is_vip=is_vip)

        income_per_item = (
            round(rental["income"] / rental["count"], 8)
            if rental["count"] > 0 else None
        )

        # 5) 写入（upsert）
        row = {
            "date": today,
            "rented_count": rental["count"],
            "rented_value": round(rental["value"], 2),
            "daily_income": round(rental["income"], 2),
            "short_lease_annual": annuals["short"],
            "long_lease_annual": annuals["long"],
            "combined_annual": annuals["combined"],
            "income_per_item": income_per_item,
            "is_vip": is_vip,
            "total_inventory": total_inv,
            "inventory_value": round(inv_value, 2),
            "cost_basis": cost_basis,
            "price_change": round(price_change, 8),
        }

        stmt = sqlite_insert(DailyTracker).values([row])
        stmt = stmt.on_conflict_do_update(
            index_elements=["date"],
            set_={k: stmt.excluded[k] for k in row if k != "date"},
        )
        await db.execute(stmt)
        await db.commit()

    elapsed = _time.monotonic() - t_start
    logger.info("snapshot_daily: DONE — date=%s, income=%.2f, value=%.2f, %.1fs",
                today, rental["income"], inv_value, elapsed)
    return row



# _get_latest_prices 已提取到 app/services/pricing.py，通过 import 引入


# ══════════════════════════════════════════════════════════════
#  CRUD
# ══════════════════════════════════════════════════════════════
#  配置（成本基准等）
# ══════════════════════════════════════════════════════════════

_DEFAULT_COST_BASIS = 4400000.0  # 默认成本基准 440万


async def _get_cost_basis(db: AsyncSession) -> float:
    """从 tracker_config 读取成本基准"""
    row = (await db.execute(
        select(TrackerConfig.value).where(TrackerConfig.key == "cost_basis")
    )).scalar()
    if row:
        try:
            return float(row)
        except (TypeError, ValueError):
            pass
    return _DEFAULT_COST_BASIS


async def get_config(db: AsyncSession) -> dict:
    """获取所有 tracker 配置"""
    rows = (await db.execute(select(TrackerConfig))).scalars().all()
    cfg = {r.key: r.value for r in rows}
    return {
        "cost_basis": float(cfg.get("cost_basis", _DEFAULT_COST_BASIS)),
    }


async def set_config(db: AsyncSession, key: str, value: str) -> dict:
    """设置 tracker 配置"""
    stmt = sqlite_insert(TrackerConfig).values([{"key": key, "value": value}])
    stmt = stmt.on_conflict_do_update(
        index_elements=["key"],
        set_={"value": stmt.excluded.value},
    )
    await db.execute(stmt)
    await db.commit()
    return await get_config(db)


# ══════════════════════════════════════════════════════════════
#  CRUD
# ══════════════════════════════════════════════════════════════

async def get_daily_records(
    db: AsyncSession,
    start: Optional[str] = None,
    end: Optional[str] = None,
) -> list[dict]:
    """查询日期范围内的每日记录"""
    q = select(DailyTracker).order_by(DailyTracker.date.desc())
    if start:
        q = q.where(DailyTracker.date >= start)
    if end:
        q = q.where(DailyTracker.date <= end)

    rows = (await db.execute(q)).scalars().all()
    return [_row_to_dict(r) for r in rows]


async def get_monthly_summary(db: AsyncSession, year: int, month: int) -> dict:
    """动态聚合月度汇总，缺失天数按平均值填充预估"""
    days_in_month = monthrange(year, month)[1]
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year:04d}-{month:02d}-{days_in_month:02d}"

    rows = (await db.execute(
        select(DailyTracker)
        .where(DailyTracker.date >= start, DailyTracker.date <= end)
        .order_by(DailyTracker.date.asc())
    )).scalars().all()

    if not rows:
        return {"year": year, "month": month, "days": 0}

    recorded_days = len(rows)
    incomes = [r.daily_income for r in rows if r.daily_income is not None]
    total_income = sum(incomes)

    # 按每天的 is_vip 动态计算服务费（VIP=10%, 非VIP=20%）
    service_fee = 0.0
    net_rental = 0.0
    for r in rows:
        if r.daily_income is not None:
            rate = _VIP_FEE_RATE if r.is_vip else _FEE_RATE
            service_fee += r.daily_income * rate
            net_rental += r.daily_income * (1 - rate)
    service_fee = round(service_fee, 2)
    net_rental = round(net_rental, 2)

    # 缺失天数预估：用已有数据的日均值填充
    is_estimated = recorded_days < days_in_month
    if is_estimated and recorded_days > 0:
        avg_daily_income = total_income / recorded_days
        avg_daily_net = net_rental / recorded_days
        avg_daily_fee = service_fee / recorded_days
        est_total_income = round(avg_daily_income * days_in_month, 2)
        est_service_fee = round(avg_daily_fee * days_in_month, 2)
        est_net_rental = round(avg_daily_net * days_in_month, 2)
    else:
        est_total_income = round(total_income, 2)
        est_service_fee = service_fee
        est_net_rental = net_rental

    # 月末库存价值（取最后一天有值的记录）
    last_val = None
    for r in reversed(rows):
        if r.inventory_value and r.inventory_value > 0:
            last_val = r.inventory_value
            break

    net_annual = round(est_net_rental * 12 / last_val, 6) if last_val else None

    # 综合年化平均值
    annuals = [r.combined_annual for r in rows if r.combined_annual is not None]
    avg_annual = round(sum(annuals) / len(annuals), 6) if annuals else None

    return {
        "year": year,
        "month": month,
        "days": recorded_days,
        "days_in_month": days_in_month,
        "is_estimated": is_estimated,
        "total_income": round(total_income, 2),
        "service_fee": service_fee,
        "net_rental": net_rental,
        "est_total_income": est_total_income,
        "est_service_fee": est_service_fee,
        "est_net_rental": est_net_rental,
        "net_rental_annual": net_annual,
        "avg_combined_annual": avg_annual,
        "last_inventory_value": last_val,
    }


async def update_record(db: AsyncSession, date_str: str, fields: dict) -> dict | None:
    """手动编辑某天的数据"""
    row = (await db.execute(
        select(DailyTracker).where(DailyTracker.date == date_str)
    )).scalar_one_or_none()

    if not row:
        return None

    allowed = {"steamdt_index", "notes", "rented_count", "rented_value",
               "daily_income", "total_inventory", "inventory_value", "is_vip", "cost_basis"}
    for k, v in fields.items():
        if k in allowed:
            setattr(row, k, v)

    # 如果修改了核心字段或 VIP 状态，重新计算年化
    if any(k in fields for k in ("daily_income", "rented_value", "rented_count", "is_vip")):
        vip = row.is_vip if hasattr(row, 'is_vip') and row.is_vip is not None else True
        annuals = _calc_annuals(row.daily_income or 0, row.rented_value or 0, is_vip=vip)
        row.short_lease_annual = annuals["short"]
        row.long_lease_annual = annuals["long"]
        row.combined_annual = annuals["combined"]
        if row.rented_count and row.daily_income:
            row.income_per_item = round(row.daily_income / row.rented_count, 8)

    # 如果修改了成本基准或库存价值，重新计算涨跌
    if any(k in fields for k in ("cost_basis", "inventory_value")):
        cb = row.cost_basis or _DEFAULT_COST_BASIS
        iv = row.inventory_value or 0
        row.price_change = round((iv - cb) / cb, 8) if cb > 0 else 0.0

    await db.commit()
    return _row_to_dict(row)


# ══════════════════════════════════════════════════════════════
#  Excel 导入 / 导出
# ══════════════════════════════════════════════════════════════

async def import_from_excel(db: AsyncSession, file_bytes: bytes) -> dict:
    """
    从 xlsx 导入每日数据。
    期望列顺序与 Excel KHH sheet 一致：
      A=日期, B=出租件数, C=总价值, D=总收益/天,
      E=短租年化, F=长租年化, G=综合年化, H=单件收益/天,
      I=总库存, J=库存总价值, K=涨跌, L=SteamDT指数
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            skipped += 1
            continue

        raw_date = row[0]
        if raw_date is None:
            skipped += 1
            continue

        # 解析日期
        if isinstance(raw_date, datetime):
            d = raw_date.strftime("%Y-%m-%d")
        elif isinstance(raw_date, date):
            d = raw_date.isoformat()
        elif isinstance(raw_date, (int, float)):
            # Excel serial date
            from datetime import timedelta
            d = (datetime(1899, 12, 30) + timedelta(days=int(raw_date))).strftime("%Y-%m-%d")
        else:
            d = str(raw_date).strip()[:10]

        rented_count = _safe_int(row[1]) if len(row) > 1 else None
        rented_value = _safe_float(row[2]) if len(row) > 2 else None
        daily_income = _safe_float(row[3]) if len(row) > 3 else None

        if rented_count is None and daily_income is None:
            skipped += 1
            continue

        rec = {
            "date": d,
            "rented_count": rented_count,
            "rented_value": rented_value,
            "daily_income": daily_income,
            "short_lease_annual": _safe_float(row[4]) if len(row) > 4 else None,
            "long_lease_annual": _safe_float(row[5]) if len(row) > 5 else None,
            "combined_annual": _safe_float(row[6]) if len(row) > 6 else None,
            "income_per_item": _safe_float(row[7]) if len(row) > 7 else None,
            "total_inventory": _safe_int(row[8]) if len(row) > 8 else None,
            "inventory_value": _safe_float(row[9]) if len(row) > 9 else None,
            "price_change": _safe_float(row[10]) if len(row) > 10 else None,
            "steamdt_index": _safe_float(row[11]) if len(row) > 11 else None,
        }

        stmt = sqlite_insert(DailyTracker).values([rec])
        stmt = stmt.on_conflict_do_update(
            index_elements=["date"],
            set_={k: stmt.excluded[k] for k in rec if k != "date" and rec[k] is not None},
        )
        await db.execute(stmt)
        imported += 1

    await db.commit()
    return {"imported": imported, "skipped": skipped}


async def export_to_excel(db: AsyncSession, start: Optional[str] = None, end: Optional[str] = None) -> bytes:
    """导出每日数据为 xlsx"""
    import openpyxl

    records = await get_daily_records(db, start, end)
    # 按日期正序
    records.sort(key=lambda r: r["date"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Tracker"

    headers = [
        "日期", "出租总件数", "总价值", "总收益/天",
        "短租年化", "长租年化", "综合年化", "单件收益/天",
        "总库存", "库存总价值", "涨跌", "SteamDT指数", "备注",
    ]
    ws.append(headers)

    fields = [
        "date", "rented_count", "rented_value", "daily_income",
        "short_lease_annual", "long_lease_annual", "combined_annual", "income_per_item",
        "total_inventory", "inventory_value", "price_change", "steamdt_index", "notes",
    ]
    for rec in records:
        ws.append([rec.get(f) for f in fields])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════════════

def _row_to_dict(r: DailyTracker) -> dict:
    # created_at 格式化为 ISO8601 UTC 字符串
    created_at_str = None
    if r.created_at:
        if r.created_at.tzinfo is None:
            created_at_str = r.created_at.replace(tzinfo=timezone.utc).isoformat()
        else:
            created_at_str = r.created_at.isoformat()

    return {
        "date": r.date,
        "created_at": created_at_str,
        "rented_count": r.rented_count,
        "rented_value": r.rented_value,
        "daily_income": r.daily_income,
        "short_lease_annual": r.short_lease_annual,
        "long_lease_annual": r.long_lease_annual,
        "combined_annual": r.combined_annual,
        "income_per_item": r.income_per_item,
        "total_inventory": r.total_inventory,
        "inventory_value": r.inventory_value,
        "price_change": r.price_change,
        "cost_basis": r.cost_basis,
        "steamdt_index": r.steamdt_index,
        "is_vip": r.is_vip,
        "notes": r.notes,
    }


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        f = float(v)
        return f if not (f != f) else None  # NaN check
    except (TypeError, ValueError):
        return None


def _safe_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None
