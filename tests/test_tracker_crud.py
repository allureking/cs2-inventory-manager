"""
tracker.py CRUD / 聚合 / Excel 测试（内存库）。

覆盖：
  - get_config / set_config（kv，默认成本基准回退）
  - get_daily_records（空 / 日期范围过滤 / 倒序）
  - get_monthly_summary（VIP 10% vs 非VIP 20% 动态费率、缺天预估、月末库存价值、空月）
  - update_record（不存在→None、字段白名单、改核心字段重算年化、改成本/库存重算涨跌）
  - import_from_excel / export_to_excel（往返）
"""

import asyncio
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from app.services import tracker
from app.services.tracker import (
    get_config,
    set_config,
    get_daily_records,
    get_monthly_summary,
    update_record,
    import_from_excel,
    export_to_excel,
    _DEFAULT_COST_BASIS,
    _VIP_FEE_RATE,
    _FEE_RATE,
)
from app.models.db_models import DailyTracker
from tests.conftest import memory_db


def _run(coro):
    return asyncio.run(coro)


async def _add_days(Session, rows):
    async with Session() as db:
        for r in rows:
            db.add(DailyTracker(**r))
        await db.commit()


def day(date, **kw):
    d = dict(date=date)
    d.update(kw)
    return d


# ── config ─────────────────────────────────────────────────────────────────


def test_get_config_default():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                cfg = await get_config(db)
                assert cfg["cost_basis"] == _DEFAULT_COST_BASIS
    _run(body())


def test_set_then_get_config():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                await set_config(db, "cost_basis", "5000000")
                cfg = await get_config(db)
                assert cfg["cost_basis"] == 5000000.0
    _run(body())


def test_set_config_upsert_overwrites():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                await set_config(db, "cost_basis", "1")
                await set_config(db, "cost_basis", "2")
                cfg = await get_config(db)
                assert cfg["cost_basis"] == 2.0
    _run(body())


# ── get_daily_records ──────────────────────────────────────────────────────


def test_daily_records_empty():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                assert await get_daily_records(db) == []
    _run(body())


def test_daily_records_desc_order():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", daily_income=10.0),
                day("2026-06-03", daily_income=30.0),
                day("2026-06-02", daily_income=20.0),
            ])
            async with Session() as db:
                recs = await get_daily_records(db)
                assert [r["date"] for r in recs] == ["2026-06-03", "2026-06-02", "2026-06-01"]
    _run(body())


def test_daily_records_range_filter():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-05-31"), day("2026-06-01"), day("2026-06-02"), day("2026-06-03"),
            ])
            async with Session() as db:
                recs = await get_daily_records(db, start="2026-06-01", end="2026-06-02")
                assert [r["date"] for r in recs] == ["2026-06-02", "2026-06-01"]
    _run(body())


# ── get_monthly_summary ────────────────────────────────────────────────────


def test_monthly_empty():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 6)
                assert r == {"year": 2026, "month": 6, "days": 0}
    _run(body())


def test_monthly_vip_vs_nonvip_fee():
    # day1 VIP(费率10%), day2 非VIP(费率20%), 各 income=1000
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", daily_income=1000.0, is_vip=True, inventory_value=100.0),
                day("2026-06-02", daily_income=1000.0, is_vip=False, inventory_value=200.0),
            ])
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 6)
                # service_fee = 1000*0.1 + 1000*0.2 = 300
                assert r["service_fee"] == pytest.approx(300.0)
                # net_rental = 1000*0.9 + 1000*0.8 = 1700
                assert r["net_rental"] == pytest.approx(1700.0)
                assert r["total_income"] == pytest.approx(2000.0)
                assert r["days"] == 2
    _run(body())


def test_monthly_estimation_partial_month():
    # 6 月 30 天,只录 2 天 → is_estimated=True, est = 日均 * 30
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", daily_income=1000.0, is_vip=True, inventory_value=100.0),
                day("2026-06-02", daily_income=1000.0, is_vip=True, inventory_value=200.0),
            ])
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 6)
                assert r["is_estimated"] is True
                assert r["days_in_month"] == 30
                # est_total_income = (2000/2) * 30 = 30000
                assert r["est_total_income"] == pytest.approx(30000.0)
                # 两天都 VIP：fee=200, net=1800 → est_fee=(200/2)*30=3000, est_net=(1800/2)*30=27000
                assert r["est_service_fee"] == pytest.approx(3000.0)
                assert r["est_net_rental"] == pytest.approx(27000.0)
    _run(body())


def test_monthly_last_inventory_value_picks_last_positive():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", daily_income=10.0, is_vip=True, inventory_value=111.0),
                day("2026-06-02", daily_income=10.0, is_vip=True, inventory_value=222.0),
                day("2026-06-03", daily_income=10.0, is_vip=True, inventory_value=0.0),  # 0 跳过
            ])
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 6)
                assert r["last_inventory_value"] == 222.0
    _run(body())


def test_monthly_full_month_not_estimated():
    # 2026-02 共 28 天,全部录入 → is_estimated=False, est==actual
    async def body():
        async with memory_db() as Session:
            rows = [day(f"2026-02-{d:02d}", daily_income=100.0, is_vip=True, inventory_value=1000.0)
                    for d in range(1, 29)]
            await _add_days(Session, rows)
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 2)
                assert r["days"] == 28
                assert r["days_in_month"] == 28
                assert r["is_estimated"] is False
                assert r["total_income"] == pytest.approx(2800.0)
                assert r["est_total_income"] == pytest.approx(r["total_income"])
                assert r["est_net_rental"] == pytest.approx(r["net_rental"])
    _run(body())


def test_monthly_avg_combined_annual():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", daily_income=10.0, is_vip=True, combined_annual=0.10, inventory_value=1.0),
                day("2026-06-02", daily_income=10.0, is_vip=True, combined_annual=0.20, inventory_value=1.0),
            ])
            async with Session() as db:
                r = await get_monthly_summary(db, 2026, 6)
                assert r["avg_combined_annual"] == pytest.approx(0.15)
    _run(body())


# ── update_record ──────────────────────────────────────────────────────────


def test_update_nonexistent_returns_none():
    async def body():
        async with memory_db() as Session:
            async with Session() as db:
                assert await update_record(db, "2099-01-01", {"daily_income": 5}) is None
    _run(body())


def test_update_recalcs_annuals():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [day("2026-06-01", daily_income=0.0, rented_value=0.0, is_vip=True)])
            async with Session() as db:
                out = await update_record(db, "2026-06-01",
                                          {"daily_income": 1000.0, "rented_value": 1_000_000.0})
                # 改了核心字段 → 年化被重算（从 None 变为有值）
                assert out["combined_annual"] is not None
                assert out["short_lease_annual"] is not None
    _run(body())


def test_update_recalcs_price_change():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [day("2026-06-01", daily_income=1.0)])
            async with Session() as db:
                out = await update_record(db, "2026-06-01",
                                          {"cost_basis": 1000.0, "inventory_value": 1500.0})
                # price_change = (1500-1000)/1000 = 0.5
                assert out["price_change"] == pytest.approx(0.5)
    _run(body())


def test_update_ignores_unlisted_fields():
    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [day("2026-06-01", daily_income=1.0)])
            async with Session() as db:
                out = await update_record(db, "2026-06-01", {"date": "hacked", "notes": "ok"})
                # date 不在白名单 → 不变；notes 在白名单 → 更新
                assert out["date"] == "2026-06-01"
                assert out["notes"] == "ok"
    _run(body())


# ── Excel 往返 ─────────────────────────────────────────────────────────────


def test_import_from_excel():
    import openpyxl

    async def body():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["日期", "件数", "价值", "收益"])  # 表头(被跳过)
        ws.append(["2026-06-01", 100, 5000.0, 250.0])
        ws.append(["2026-06-02", 110, 5500.0, 275.0])
        buf = io.BytesIO()
        wb.save(buf)

        async with memory_db() as Session:
            async with Session() as db:
                res = await import_from_excel(db, buf.getvalue())
                assert res["imported"] == 2
                recs = await get_daily_records(db)
                by_date = {r["date"]: r for r in recs}
                assert by_date["2026-06-01"]["rented_count"] == 100
                assert by_date["2026-06-02"]["daily_income"] == 275.0
    _run(body())


def test_import_skips_blank_rows():
    import openpyxl

    async def body():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["日期", "件数", "价值", "收益"])
        ws.append([None, None, None, None])          # 空行跳过
        ws.append(["2026-06-01", 100, 5000.0, 250.0])
        buf = io.BytesIO()
        wb.save(buf)
        async with memory_db() as Session:
            async with Session() as db:
                res = await import_from_excel(db, buf.getvalue())
                assert res["imported"] == 1
                assert res["skipped"] >= 1
    _run(body())


def test_export_to_excel_roundtrip():
    import openpyxl

    async def body():
        async with memory_db() as Session:
            await _add_days(Session, [
                day("2026-06-01", rented_count=100, daily_income=250.0),
            ])
            async with Session() as db:
                data = await export_to_excel(db)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "日期"  # 表头
        assert rows[1][0] == "2026-06-01"
    _run(body())


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
