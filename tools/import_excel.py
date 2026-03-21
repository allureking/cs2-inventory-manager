#!/usr/bin/env python3
"""
一次性脚本：从 csgo饰品收益.xlsx 的 KHH sheet 导入历史数据到 daily_tracker 表。
仅导入第 226 行之后的数据（2025年7月底建仓后）。

用法: python tools/import_excel.py
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import asyncio
from datetime import datetime, timedelta

import openpyxl
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from app.core.database import AsyncSessionLocal, init_db
from app.models.db_models import DailyTracker


def _safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def _safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # Excel serial date number
        return (datetime(1899, 12, 30) + timedelta(days=int(v))).strftime("%Y-%m-%d")
    return None


async def main():
    xlsx_path = Path(__file__).resolve().parent.parent / "csgo饰品收益.xlsx"
    if not xlsx_path.exists():
        print(f"❌ 文件不存在: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["KHH"]

    await init_db()

    records = []
    start_row = 226  # 建仓后数据起始行

    for row_num in range(start_row, ws.max_row + 1):
        raw_date = ws.cell(row=row_num, column=1).value
        rented_count = ws.cell(row=row_num, column=2).value
        daily_income = ws.cell(row=row_num, column=4).value

        # 跳过无数据行
        if raw_date is None or (rented_count is None and daily_income is None):
            continue

        d = _parse_date(raw_date)
        if not d:
            continue

        rec = {
            "date": d,
            "rented_count": _safe_int(rented_count),
            "rented_value": _safe_float(ws.cell(row=row_num, column=3).value),
            "daily_income": _safe_float(daily_income),
            "short_lease_annual": _safe_float(ws.cell(row=row_num, column=5).value),
            "long_lease_annual": _safe_float(ws.cell(row=row_num, column=6).value),
            "combined_annual": _safe_float(ws.cell(row=row_num, column=7).value),
            "income_per_item": _safe_float(ws.cell(row=row_num, column=8).value),
            "total_inventory": _safe_int(ws.cell(row=row_num, column=9).value),
            "inventory_value": _safe_float(ws.cell(row=row_num, column=10).value),
            "price_change": _safe_float(ws.cell(row=row_num, column=11).value),
            "steamdt_index": _safe_float(ws.cell(row=row_num, column=12).value),
        }
        records.append(rec)

    print(f"📊 解析到 {len(records)} 条记录 (行 {start_row}-{ws.max_row})")

    if not records:
        print("❌ 无数据可导入")
        return

    # 打印前3条和后3条预览
    print("\n前3条:")
    for r in records[:3]:
        print(f"  {r['date']} | 出租{r['rented_count']}件 | 日租金{r['daily_income']} | 综合年化{r['combined_annual']}")
    print("后3条:")
    for r in records[-3:]:
        print(f"  {r['date']} | 出租{r['rented_count']}件 | 日租金{r['daily_income']} | 综合年化{r['combined_annual']}")

    # 写入数据库
    async with AsyncSessionLocal() as db:
        for rec in records:
            stmt = sqlite_insert(DailyTracker).values([rec])
            stmt = stmt.on_conflict_do_update(
                index_elements=["date"],
                set_={k: stmt.excluded[k] for k in rec if k != "date" and rec[k] is not None},
            )
            await db.execute(stmt)
        await db.commit()

    print(f"\n✅ 成功导入 {len(records)} 条记录到 daily_tracker 表")


if __name__ == "__main__":
    asyncio.run(main())
